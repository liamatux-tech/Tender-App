import sys
import json
import io
import requests
import urllib3
import webbrowser
import threading
from datetime import datetime
from zoneinfo import ZoneInfo
from flask import Flask, request, jsonify, render_template_string, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

BAKU_TZ = ZoneInfo('Asia/Baku')
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
}

app = Flask(__name__)

# ─────────────────────────────────────────
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ─────────────────────────────────────────

def parse_date(val):
    if not val: return None
    try:
        if isinstance(val, (int, float)) or (isinstance(val, str) and str(val).isdigit()):
            ts = float(val)
            if ts > 1e11: ts /= 1000
            return datetime.fromtimestamp(ts, BAKU_TZ).strftime('%d.%m.%Y %H:%M')
        dt = datetime.fromisoformat(str(val).replace('Z', '+00:00'))
        return dt.astimezone(BAKU_TZ).strftime('%d.%m.%Y %H:%M')
    except:
        return None

def is_expired(val):
    if not val: return False
    try:
        if isinstance(val, (int, float)) or (isinstance(val, str) and str(val).isdigit()):
            ts = float(val)
            if ts > 1e11: ts /= 1000
            return datetime.fromtimestamp(ts, BAKU_TZ) < datetime.now(BAKU_TZ)
        dt = datetime.fromisoformat(str(val).replace('Z', '+00:00'))
        return dt.astimezone(BAKU_TZ) < datetime.now(BAKU_TZ)
    except:
        return False

def get_tender(event_id):
    base = "https://etender.gov.az/api/events"
    d = requests.get(f"{base}/{event_id}", headers=HEADERS, timeout=10, verify=False).json()

    try:
        c_res = requests.get(f"{base}/{event_id}/contact-persons", headers=HEADERS, timeout=10, verify=False).json()
    except:
        c_res = []

    contact = None
    if c_res and len(c_res) > 0:
        p = c_res[0]
        contact = {
            'person': f"{p.get('fullName', '')} ({p.get('position', '')})".strip(),
            'email':  p.get('contact'),
            'phone':  p.get('phoneNumber')
        }

    bom = []
    try:
        all_items = []
        page = 1
        while True:
            bom_res = requests.get(f"{base}/{event_id}/bomLines?PageSize=1000&PageNumber={page}", headers=HEADERS, timeout=10, verify=False).json()
            items = bom_res.get('items', [])
            all_items.extend(items)
            total = bom_res.get('totalCount') or bom_res.get('total') or 0
            if len(all_items) >= total or len(items) == 0:
                break
            page += 1
        bom = [{'name': i.get('name'), 'qty': i.get('quantity'), 'unit': i.get('unitOfMeasure'), 'desc': i.get('description')} for i in all_items]
    except:
        pass

    expired = is_expired(d.get('endDate'))

    return {
        'id':      str(event_id),
        'title':   d.get('tenderName') or '—',
        'org':     d.get('organizationName') or '—',
        'voen':    d.get('organizationVoen') or '—',
        'amount':  d.get('estimatedAmount'),
        'start':   parse_date(d.get('startDate')),
        'end':     parse_date(d.get('endDate')),
        'phone':   (contact['phone'] if contact and contact['phone'] else d.get('contactPhone')) or '—',
        'email':   (contact['email'] if contact and contact['email'] else d.get('contactEmail')) or '—',
        'person':  (contact['person'] if contact and contact['person'] else 'Qeyd olunmayıb'),
        'link':    f"https://etender.gov.az/main/competition/detail/{event_id}",
        'expired': expired,
        'bom':     bom
    }

# ─────────────────────────────────────────
# МАРШРУТЫ
# ─────────────────────────────────────────

@app.route('/')
def index():
    return render_template_string(HTML_PAGE)

@app.route('/api/tender')
def api_tender():
    link = request.args.get('link', '').strip()
    if not link:
        return jsonify({'error': 'Link daxil edilməyib'}), 400

    event_id = link.rstrip('/').split('/')[-1]
    if not event_id.isdigit():
        return jsonify({'error': 'Link düzgün deyil'}), 400

    try:
        data = get_tender(event_id)
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export')
def api_export():
    link = request.args.get('link', '').strip()
    if not link: return jsonify({'error': 'Link yoxdur'}), 400
    event_id = link.rstrip('/').split('/')[-1]
    try:
        d = get_tender(event_id)
        wb = build_excel(d)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name=f"tender_{event_id}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def build_excel(d):
    wb = Workbook()
    C_DARK, C_ACCENT, C_LIGHT, C_WHITE, C_GREEN, C_RED, C_GRAY, C_MUTED, C_BORDER = "0D1117", "00B4CC", "E8F7F9", "FFFFFF", "0E7C59", "C0392B", "F2F4F5", "6B7280", "D1D5DB"
    thin = Side(style='thin', color=C_BORDER)
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hfill(hex_color): return PatternFill("solid", fgColor=hex_color)

    # ШИТ 1
    ws = wb.active
    ws.title = "Tender Məlumatı"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 75 # Фиксированная большая ширина для данных

    ws.merge_cells("A1:B1")
    ws["A1"] = "TENDER MƏLUMAT KARTI"
    ws["A1"].font = Font(bold=True, size=16, color=C_WHITE); ws["A1"].fill = hfill(C_DARK); ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    fields = [
        ("Tenderin adı", d['title'], True),
        ("Təşkilat", d['org'], False),
        ("VÖEN", d['voen'], False),
        ("Məbləğ (AZN)", f"{float(d['amount']):,.2f} ₼" if d['amount'] else "—", False),
        ("Bitmə tarixi", d['end'] or "—", False),
        ("Əlaqədar şəxs", d['person'], False),
        ("Telefon", d['phone'], False),
        ("Email", d['email'], False),
        ("Link", d['link'], False),
    ]

    for i, (label, val, wrap) in enumerate(fields):
        row = 5 + i
        ws.cell(row=row, column=1, value=label).font = Font(bold=True); ws.cell(row=row, column=1).fill = hfill(C_GRAY); ws.cell(row=row, column=1).border = border_all
        c = ws.cell(row=row, column=2, value=val)
        c.border = border_all; c.alignment = Alignment(wrap_text=wrap)
        if label == "Məbləğ (AZN)": c.font = Font(bold=True, color=C_GREEN)

    # ШИТ 2: BOM
    ws2 = wb.create_sheet("Spesifikasiya")
    ws2.sheet_view.showGridLines = False
    headers = ["№", "Məhsul / Xidmət", "Açıqlama", "Miqdar", "Ölçü vahidi"]
    widths = [6, 45, 50, 12, 15] # Улучшенная ширина колонок
    
    for ci, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    for ci, title in enumerate(headers, 1):
        c = ws2.cell(row=1, column=ci, value=title)
        c.font = Font(bold=True, color=C_WHITE); c.fill = hfill("1E293B"); c.alignment = Alignment(horizontal="center")

    if d['bom']:
        for i, item in enumerate(d['bom']):
            row = 2 + i
            data = [i+1, item['name'], item.get('desc'), item['qty'], item['unit']]
            for ci, val in enumerate(data, 1):
                c = ws2.cell(row=row, column=ci, value=val)
                c.border = border_all
                c.alignment = Alignment(wrap_text=True, vertical="top")
    
    return wb

HTML_PAGE = """<!DOCTYPE html>
<html lang="az">
<head>
<meta charset="UTF-8">
<title>Tender Axtarış</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=Syne:wght@400;700;800&display=swap" rel="stylesheet">
<style>
  :root { --bg: #0b0e14; --surface: #12161f; --border: #1e2535; --accent: #00e5ff; --text: #e2e8f0; --muted: #4a5568; --success: #10b981; --danger: #ef4444; }
  body { background: var(--bg); color: var(--text); font-family: 'Syne', sans-serif; padding: 40px 20px; }
  .wrapper { max-width: 1000px; margin: 0 auto; }
  .search-box { background: var(--surface); border: 1px solid var(--border); padding: 15px; border-radius: 8px; display: flex; gap: 10px; margin-bottom: 30px; }
  input { flex: 1; background: transparent; border: none; color: #fff; outline: none; font-family: 'DM Mono'; }
  button { background: var(--accent); color: #000; border: none; padding: 10px 25px; border-radius: 5px; font-weight: 700; cursor: pointer; }
  .info-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 10px; margin-bottom: 30px; }
  .info-cell { background: var(--surface); padding: 15px; border: 1px solid var(--border); border-radius: 6px; }
  .full { grid-column: 1/-1; }
  .label { font-family: 'DM Mono'; font-size: 11px; color: var(--muted); text-transform: uppercase; }
  .value { margin-top: 5px; font-weight: 600; }
  table { width: 100%; border-collapse: collapse; margin-top: 20px; background: var(--surface); border-radius: 8px; overflow: hidden; }
  th { background: #0d1017; color: var(--muted); font-family: 'DM Mono'; font-size: 11px; text-align: left; padding: 12px; }
  td { padding: 12px; border-top: 1px solid var(--border); font-size: 13px; vertical-align: top; }
  .desc-col { color: #94a3b8; font-size: 12px; max-width: 300px; }
  #excelBtn { background: var(--success); color: #fff; margin-bottom: 15px; display: none; }
</style>
</head>
<body>
<div class="wrapper">
  <h1>Tender <span>Axtarış</span></h1>
  <div class="search-box">
    <input type="text" id="linkInput" placeholder="etender.gov.az linkini bura yapışdırın...">
    <button onclick="fetchTender()">Axtar</button>
  </div>
  <button id="excelBtn" onclick="downloadExcel()">Excel Yüklə</button>
  <div id="result"></div>
</div>
<script>
  async function fetchTender() {
    const link = document.getElementById('linkInput').value;
    const res = await fetch('/api/tender?link=' + encodeURIComponent(link));
    const d = await res.json();
    if(d.error) return alert(d.error);
    
    let bomRows = d.bom.map((b, i) => `
      <tr>
        <td>${i+1}</td>
        <td><b>${b.name||'—'}</b></td>
        <td class="desc-col">${b.desc||'—'}</td>
        <td style="color:var(--accent)">${b.qty||'—'}</td>
        <td>${b.unit||'—'}</td>
      </tr>`).join('');

    document.getElementById('result').innerHTML = `
      <div class="info-grid">
        <div class="info-cell full"><div class="label">Tenderin Adı</div><div class="value">${d.title}</div></div>
        <div class="info-cell"><div class="label">Təşkilat</div><div class="value">${d.org}</div></div>
        <div class="info-cell"><div class="label">Məbləğ</div><div class="value" style="color:var(--success)">${Number(d.amount).toLocaleString()} AZN</div></div>
      </div>
      <table>
        <thead><tr><th>#</th><th>Məhsul</th><th>Açıqlama (Description)</th><th>Miqdar</th><th>Vahid</th></tr></thead>
        <tbody>${bomRows || '<tr><td colspan="5">Spesifikasiya tapılmadı</td></tr>'}</tbody>
      </table>`;
    document.getElementById('excelBtn').style.display = 'block';
  }
  function downloadExcel() {
    window.location.href = '/api/export?link=' + encodeURIComponent(document.getElementById('linkInput').value);
  }
</script>
</body>
</html>"""

def open_browser():
    import time
    time.sleep(1)
    webbrowser.open('http://127.0.0.1:5000')

if __name__ == '__main__':
    threading.Thread(target=open_browser, daemon=True).start()
    app.run(host='0.0.0.0', debug=False, port=5000)