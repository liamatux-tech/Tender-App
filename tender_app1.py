import sys
import json
import io
import requests
import urllib3
import webbrowser
import threading
from datetime import datetime
from zoneinfo import ZoneInfo
from flask import Flask, request, jsonify, render_template_string, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

BAKU_TZ = ZoneInfo('Asia/Baku')
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
}

app = Flask(__name__)

# ─────────────────────────────────────────
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ─────────────────────────────────────────

def parse_date(val):
    if not val: return None
    try:
        if isinstance(val, (int, float)) or (isinstance(val, str) and str(val).isdigit()):
            ts = float(val)
            if ts > 1e11: ts /= 1000
            return datetime.fromtimestamp(ts, BAKU_TZ).strftime('%d.%m.%Y %H:%M')
        dt = datetime.fromisoformat(str(val).replace('Z', '+00:00'))
        return dt.astimezone(BAKU_TZ).strftime('%d.%m.%Y %H:%M')
    except:
        return None

def is_expired(val):
    if not val: return False
    try:
        if isinstance(val, (int, float)) or (isinstance(val, str) and str(val).isdigit()):
            ts = float(val)
            if ts > 1e11: ts /= 1000
            return datetime.fromtimestamp(ts, BAKU_TZ) < datetime.now(BAKU_TZ)
        dt = datetime.fromisoformat(str(val).replace('Z', '+00:00'))
        return dt.astimezone(BAKU_TZ) < datetime.now(BAKU_TZ)
    except:
        return False

def get_tender(event_id):
    base = "https://etender.gov.az/api/events"

    d = requests.get(f"{base}/{event_id}", headers=HEADERS, timeout=10, verify=False).json()

    try:
        c_res = requests.get(f"{base}/{event_id}/contact-persons", headers=HEADERS, timeout=10, verify=False).json()
    except:
        c_res = []

    contact = None
    if c_res and len(c_res) > 0:
        p = c_res[0]
        contact = {
            'person': f"{p.get('fullName', '')} ({p.get('position', '')})".strip(),
            'email':  p.get('contact'),
            'phone':  p.get('phoneNumber')
        }

    try:
      all_items = []
      page = 1
      while True:
          bom_res = requests.get(f"{base}/{event_id}/bomLines?PageSize=1000&PageNumber={page}", headers=HEADERS, timeout=10, verify=False).json()
          items = bom_res.get('items', [])
          all_items.extend(items)
          total = bom_res.get('totalCount') or bom_res.get('total') or 0
          if len(all_items) >= total or len(items) == 0:
              break
          page += 1
      bom = [{'name': i.get('name'), 'qty': i.get('quantity'), 'unit': i.get('unitOfMeasure'), 'desc': i.get('description')} for i in bom_res.get('items', [])]
    except:
      bom = []

    expired = is_expired(d.get('endDate'))

    return {
        'id':      str(event_id),
        'title':   d.get('tenderName') or '—',
        'org':     d.get('organizationName') or '—',
        'voen':    d.get('organizationVoen') or '—',
        'amount':  d.get('estimatedAmount'),
        'start':   parse_date(d.get('startDate')),
        'end':     parse_date(d.get('endDate')),
        'phone':   (contact['phone'] if contact and contact['phone'] else d.get('contactPhone')) or '—',
        'email':   (contact['email'] if contact and contact['email'] else d.get('contactEmail')) or '—',
        'person':  (contact['person'] if contact and contact['person'] else 'Qeyd olunmayıb'),
        'link':    f"https://etender.gov.az/main/competition/detail/{event_id}",
        'expired': expired,
        'bom':     bom
    }

# ─────────────────────────────────────────
# МАРШРУТЫ
# ─────────────────────────────────────────

@app.route('/')
def index():
    return render_template_string(HTML_PAGE)

@app.route('/api/tender')
def api_tender():
    link = request.args.get('link', '').strip()
    if not link:
        return jsonify({'error': 'Link daxil edilməyib'}), 400

    event_id = link.rstrip('/').split('/')[-1]
    if not event_id.isdigit():
        return jsonify({'error': 'Link düzgün deyil'}), 400

    try:
        data = get_tender(event_id)
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export')
def api_export():
    link = request.args.get('link', '').strip()
    if not link:
        return jsonify({'error': 'Link yoxdur'}), 400
    event_id = link.rstrip('/').split('/')[-1]
    if not event_id.isdigit():
        return jsonify({'error': 'Link düzgün deyil'}), 400
    try:
        d = get_tender(event_id)
        wb = build_excel(d)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name=f"tender_{event_id}.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def build_excel(d):
    wb = Workbook()

    C_DARK   = "0D1117"
    C_ACCENT = "00B4CC"
    C_LIGHT  = "E8F7F9"
    C_WHITE  = "FFFFFF"
    C_GREEN  = "0E7C59"
    C_RED    = "C0392B"
    C_GRAY   = "F2F4F5"
    C_MUTED  = "6B7280"
    C_BORDER = "D1D5DB"

    thin  = Side(style='thin',   color=C_BORDER)
    thick = Side(style='medium', color=C_ACCENT)
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hfill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    # ── ШИТ 1: ОБЩАЯ ИНФОРМАЦИЯ ───────────────────────────────────────────
    ws = wb.active
    ws.title = "Tender Məlumatı"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 58

    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 48
    c = ws["A1"]
    c.value = "TENDER MƏLUMAT KARTI"
    c.font = Font(bold=True, size=16, color=C_WHITE, name="Arial")
    c.fill = hfill(C_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:B2")
    c = ws["A2"]
    c.value = f"etender.gov.az  |  ID: {d['id']}  |  {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    c.font = Font(size=9, color=C_ACCENT, name="Arial")
    c.fill = hfill(C_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    for col in ['A', 'B']:
        ws[f"{col}3"].fill = hfill(C_ACCENT)
    ws.row_dimensions[3].height = 5

    ws.merge_cells("A4:B4")
    status_text  = "✕  MÜDDƏTI BİTMİŞDİR" if d['expired'] else "●  AKTİV TENDER"
    status_color = C_RED if d['expired'] else C_GREEN
    c = ws["A4"]
    c.value = status_text
    c.font = Font(bold=True, size=11, color=C_WHITE, name="Arial")
    c.fill = hfill(status_color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 6

    fields = [
        ("Tenderin adı",   d['title'],  True),
        ("Təşkilat",       d['org'],    False),
        ("VÖEN",           d['voen'],   False),
        ("Məbləğ (AZN)",   f"{float(d['amount']):,.2f} ₼" if d['amount'] else "—", False),
        ("Başlama tarixi", d['start'] or "—", False),
        ("Bitmə tarixi",   d['end']   or "—", False),
        ("Əlaqədar şəxs",  d['person'], False),
        ("Telefon",        d['phone'],  False),
        ("Email",          d['email'],  False),
        ("Link",           d['link'],   False),
    ]

    for i, (label, value, wrap) in enumerate(fields):
        row = 6 + i
        ws.row_dimensions[row].height = 32 if wrap else 22
        bg_l = C_GRAY if i % 2 == 0 else C_WHITE
        bg_v = C_LIGHT if i % 2 == 0 else C_WHITE

        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(bold=True, size=10, color="374151", name="Arial")
        lc.fill = hfill(bg_l)
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border = border_all

        vc = ws.cell(row=row, column=2, value=value)
        vc.font = Font(size=10, color="111827", name="Arial")
        vc.fill = hfill(bg_v)
        vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)
        vc.border = border_all

        if label == "Məbləğ (AZN)":
            vc.font = Font(bold=True, size=12, color=C_GREEN, name="Arial")
        if label == "Link":
            vc.font = Font(size=10, color=C_ACCENT, name="Arial", underline="single")

    sep = 6 + len(fields)
    ws.row_dimensions[sep].height = 10
    ws.merge_cells(f"A{sep+1}:B{sep+1}")
    c = ws.cell(row=sep+1, column=1)
    c.value = f"Fayl yaradılma tarixi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    c.font = Font(size=8, color=C_MUTED, name="Arial", italic=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[sep+1].height = 18

    # ── ШИТ 2: BOM ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Spesifikasiya")
    ws2.sheet_view.showGridLines = False
    for ci, w in enumerate([8, 36, 36, 14, 16], 1):
      ws2.column_dimensions[get_column_letter(ci)].width = w

    ws2.merge_cells("A1:D1")
    c = ws2["A1"]
    c.value = "SPESİFİKASİYA / MƏHSUL SİYAHISI"
    c.font = Font(bold=True, size=14, color=C_WHITE, name="Arial")
    c.fill = hfill(C_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 40

    ws2.merge_cells("A2:D2")
    c = ws2["A2"]
    c.value = f"{d['title']}  |  ID: {d['id']}"
    c.font = Font(size=9, color=C_ACCENT, name="Arial")
    c.fill = hfill(C_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[2].height = 18

    for col in range(1, 5):
        ws2.cell(row=3, column=col).fill = hfill(C_ACCENT)
    ws2.row_dimensions[3].height = 5

    for ci, title in enumerate(["№", "Məhsul / Xidmət", "Açıqlama", "Miqdar", "Ölçü vahidi"], 1):
        c = ws2.cell(row=4, column=ci, value=title)
        c.font = Font(bold=True, size=10, color=C_WHITE, name="Arial")
        c.fill = hfill("1E293B")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_all
    ws2.row_dimensions[4].height = 24

    if d['bom']:
        for i, item in enumerate(d['bom']):
            row = 5 + i
            ws2.row_dimensions[row].height = 20
            fill = hfill(C_LIGHT if i % 2 == 0 else C_WHITE)
            cells_data = [
              (1, i+1, Alignment(horizontal="center", vertical="center"), Font(size=9, color=C_MUTED, name="Arial")),
              (2, item['name'] or "—", Alignment(horizontal="left", vertical="center", wrap_text=True), Font(size=10, color="111827", name="Arial")),
              (3, item.get('desc') or "—", Alignment(horizontal="left", vertical="center", wrap_text=True), Font(size=9, color="374151", name="Arial")),
              (4, item['qty'], Alignment(horizontal="center", vertical="center"), Font(bold=True, size=10, color=C_ACCENT, name="Arial")),
              (5, item['unit'] or "—", Alignment(horizontal="center", vertical="center"), Font(size=10, color="374151", name="Arial")),
            ]
            
            for col, val, aln, fnt in cells_data:
                c = ws2.cell(row=row, column=col, value=val)
                c.font = fnt; c.fill = fill; c.alignment = aln; c.border = border_all

        total_row = 5 + len(d['bom'])
        ws2.row_dimensions[total_row].height = 24
        ws2.merge_cells(f"A{total_row}:B{total_row}")
        c = ws2.cell(row=total_row, column=1, value=f"Cəmi: {len(d['bom'])} mövqe")
        c.font = Font(bold=True, size=10, color=C_WHITE, name="Arial")
        c.fill = hfill(C_DARK)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border_all
        for col in range(3, 5):
            c2 = ws2.cell(row=total_row, column=col)
            c2.fill = hfill(C_DARK); c2.border = border_all
    else:
        ws2.merge_cells("A5:D5")
        c = ws2["A5"]
        c.value = "Spesifikasiya məlumatı mövcud deyil"
        c.font = Font(size=10, color=C_MUTED, name="Arial", italic=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[5].height = 30

    return wb


# ─────────────────────────────────────────
# HTML ШАБЛОН
# ─────────────────────────────────────────

HTML_PAGE = """<!DOCTYPE html>
<html lang="az">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Tender Axtarış</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=Syne:wght@400;600;700;800&display=swap" rel="stylesheet">
<style>
  :root {
    --bg: #0b0e14;
    --surface: #12161f;
    --border: #1e2535;
    --accent: #00e5ff;
    --text: #e2e8f0;
    --muted: #4a5568;
    --success: #10b981;
    --danger: #ef4444;
  }
  * { margin:0; padding:0; box-sizing:border-box; }
  body {
    background: var(--bg);
    color: var(--text);
    font-family: 'Syne', sans-serif;
    min-height: 100vh;
    overflow-x: hidden;
  }
  body::before {
    content:'';
    position:fixed; inset:0;
    background-image:
      linear-gradient(rgba(0,229,255,0.025) 1px, transparent 1px),
      linear-gradient(90deg, rgba(0,229,255,0.025) 1px, transparent 1px);
    background-size: 40px 40px;
    pointer-events:none; z-index:0;
  }
  .wrapper { position:relative; z-index:1; max-width:860px; margin:0 auto; padding:60px 24px 80px; }

  /* Header */
  .badge {
    display:inline-flex; align-items:center; gap:6px;
    background:rgba(0,229,255,0.08); border:1px solid rgba(0,229,255,0.2);
    color:var(--accent); font-family:'DM Mono',monospace;
    font-size:11px; letter-spacing:0.1em; padding:4px 12px;
    border-radius:2px; margin-bottom:16px; text-transform:uppercase;
  }
  .badge::before {
    content:''; width:6px; height:6px;
    background:var(--accent); border-radius:50%;
    animation:pulse 2s infinite;
  }
  @keyframes pulse { 0%,100%{opacity:1;transform:scale(1)} 50%{opacity:.4;transform:scale(.8)} }
  h1 { font-size:clamp(32px,5vw,52px); font-weight:800; line-height:1.05; letter-spacing:-0.02em; color:#fff; margin-bottom:8px; }
  h1 span { color:var(--accent); }
  .subtitle { color:var(--muted); font-size:13px; font-family:'DM Mono',monospace; margin-bottom:40px; }

  /* Search */
  .search-box {
    background:var(--surface); border:1px solid var(--border);
    border-radius:6px; padding:16px 20px;
    display:flex; gap:12px; align-items:center; margin-bottom:32px;
    transition:border-color .2s;
  }
  .search-box:focus-within {
    border-color:var(--accent);
    box-shadow:0 0 0 1px rgba(0,229,255,0.12), 0 0 30px rgba(0,229,255,0.04);
  }
  #linkInput {
    flex:1; background:transparent; border:none; outline:none;
    color:var(--text); font-family:'DM Mono',monospace; font-size:13px;
  }
  #linkInput::placeholder { color:var(--muted); }
  #searchBtn {
    background:var(--accent); color:#000; border:none;
    padding:10px 24px; border-radius:4px;
    font-family:'Syne',sans-serif; font-weight:700; font-size:13px;
    cursor:pointer; letter-spacing:.05em; text-transform:uppercase;
    transition:all .2s; flex-shrink:0;
  }
  #searchBtn:hover { background:#33ecff; transform:translateY(-1px); }
  #searchBtn:disabled { opacity:.5; cursor:not-allowed; transform:none; }

  #excelBtn {
    display:none; align-items:center; gap:8px;
    background:rgba(16,185,129,0.1); border:1px solid rgba(16,185,129,0.35);
    color:var(--success); padding:10px 20px; border-radius:4px;
    font-family:'Syne',sans-serif; font-weight:700; font-size:13px;
    cursor:pointer; letter-spacing:.05em; text-transform:uppercase;
    transition:all .2s; margin-bottom:28px;
  }
  #excelBtn:hover { background:rgba(16,185,129,0.18); transform:translateY(-1px); }

  /* States */
  #loading { display:none; align-items:center; gap:12px; color:var(--muted); font-family:'DM Mono',monospace; font-size:13px; margin-bottom:24px; }
  .spinner { width:18px; height:18px; border:2px solid var(--border); border-top-color:var(--accent); border-radius:50%; animation:spin .7s linear infinite; }
  @keyframes spin { to{transform:rotate(360deg)} }
  #errorBox { display:none; background:rgba(239,68,68,0.08); border:1px solid rgba(239,68,68,0.25); color:#fca5a5; padding:14px 18px; border-radius:4px; font-family:'DM Mono',monospace; font-size:13px; margin-bottom:24px; }

  /* Result */
  #result { display:none; animation:fadeUp .4s ease; }
  @keyframes fadeUp { from{opacity:0;transform:translateY(16px)} to{opacity:1;transform:translateY(0)} }

  .result-header { display:flex; align-items:flex-start; justify-content:space-between; gap:16px; margin-bottom:24px; }
  .tender-title { font-size:clamp(16px,2.5vw,22px); font-weight:700; color:#fff; line-height:1.3; flex:1; }
  .status-badge { flex-shrink:0; font-family:'DM Mono',monospace; font-size:11px; padding:5px 12px; border-radius:2px; text-transform:uppercase; letter-spacing:.08em; }
  .status-active { background:rgba(16,185,129,0.12); border:1px solid rgba(16,185,129,0.3); color:var(--success); }
  .status-expired { background:rgba(239,68,68,0.1); border:1px solid rgba(239,68,68,0.25); color:var(--danger); }

  /* Grid */
  .info-grid { display:grid; grid-template-columns:1fr 1fr; gap:1px; background:var(--border); border:1px solid var(--border); border-radius:6px; overflow:hidden; margin-bottom:20px; }
  .info-cell { background:var(--surface); padding:18px 20px; }
  .info-cell.full { grid-column:1/-1; }
  .cell-label { font-family:'DM Mono',monospace; font-size:10px; letter-spacing:.12em; text-transform:uppercase; color:var(--muted); margin-bottom:6px; }
  .cell-value { font-size:14px; color:var(--text); font-weight:500; word-break:break-word; }
  .cell-value.mono { color:var(--accent); font-family:'DM Mono',monospace; font-size:13px; }
  .cell-value.amount { font-size:22px; font-weight:800; color:var(--success); letter-spacing:-0.02em; }
  .cell-value.expired-date { color:var(--danger); }
  .cell-value a { color:var(--accent); text-decoration:none; font-family:'DM Mono',monospace; font-size:12px; }
  .cell-value a:hover { text-decoration:underline; }

  /* BOM */
  .bom-section { margin-top:20px; }
  .section-label { font-family:'DM Mono',monospace; font-size:11px; letter-spacing:.1em; text-transform:uppercase; color:var(--muted); margin-bottom:12px; display:flex; align-items:center; gap:8px; }
  .section-label::after { content:''; flex:1; height:1px; background:var(--border); }
  .bom-count { background:rgba(0,229,255,0.1); color:var(--accent); font-size:10px; padding:2px 8px; border-radius:2px; border:1px solid rgba(0,229,255,0.2); }
  .bom-wrap { background:var(--surface); border:1px solid var(--border); border-radius:6px; overflow:hidden; }
  table { width:100%; border-collapse:collapse; font-size:13px; }
  thead th { background:#0d1017; padding:10px 14px; text-align:left; font-family:'DM Mono',monospace; font-size:10px; letter-spacing:.1em; text-transform:uppercase; color:var(--muted); border-bottom:1px solid var(--border); }
  tbody tr { border-bottom:1px solid var(--border); transition:background .15s; }
  tbody tr:last-child { border-bottom:none; }
  tbody tr:hover { background:rgba(255,255,255,0.02); }
  tbody td { padding:11px 14px; color:var(--text); }
  tbody td:first-child { color:var(--muted); font-family:'DM Mono',monospace; font-size:11px; width:36px; }
  .qty { font-family:'DM Mono',monospace; color:var(--accent) !important; }
  .no-bom { background:var(--surface); border:1px solid var(--border); border-radius:6px; padding:24px; text-align:center; color:var(--muted); font-family:'DM Mono',monospace; font-size:12px; }

  @media(max-width:560px) {
    .info-grid { grid-template-columns:1fr; }
    .result-header { flex-direction:column; }
    #searchBtn { width:100%; }
  }
</style>
</head>
<body>
<div class="wrapper">
  <header>
    <div class="badge">etender.gov.az</div>
    <h1>Tender <span>Axtarış</span></h1>
    <p class="subtitle">// tender linkini daxil et → tam məlumatı al</p>
  </header>

  <div class="search-box">
    <span style="color:var(--muted);flex-shrink:0">🔗</span>
    <input type="text" id="linkInput"
      placeholder="https://etender.gov.az/main/competition/detail/351117"
      autocomplete="off" spellcheck="false" />
    <button id="searchBtn" onclick="fetchTender()">Axtar</button>
  </div>

  <div id="loading"><div class="spinner"></div> Məlumatlar yüklənir...</div>
  <div id="errorBox"></div>
  <button id="excelBtn" onclick="downloadExcel()">⬇ Excel Yüklə</button>
  <div id="result"></div>
</div>

<script>
  document.getElementById('linkInput').addEventListener('keydown', e => {
    if (e.key === 'Enter') fetchTender();
  });

  async function fetchTender() {
    const link = document.getElementById('linkInput').value.trim();
    if (!link) return showError('Zəhmət olmasa link daxil edin.');

    setLoading(true);
    hideError();
    document.getElementById('result').style.display = 'none';
    document.getElementById('excelBtn').style.display = 'none';

    try {
      const res = await fetch('/api/tender?link=' + encodeURIComponent(link));
      const data = await res.json();
      if (!res.ok || data.error) return showError(data.error || 'Xəta baş verdi.');
      render(data);
    } catch(e) {
      showError('Serverlə əlaqə qurulmadı: ' + e.message);
    } finally {
      setLoading(false);
    }
  }

  function render(d) {
    const amt = d.amount != null ? Number(d.amount).toLocaleString('az-AZ') + ' ₼' : '—';
    const bomHTML = d.bom && d.bom.length > 0 ? `
      <div class="bom-section">
        <div class="section-label">Məhsullar / Xidmətlər <span class="bom-count">${d.bom.length} əd</span></div>
        <div class="bom-wrap">
          <table>
            <thead><tr><th>#</th><th>Məhsul / Xidmət</th><th>Açıqlama</th><th>Miqdar</th><th>Vahid</th></tr></thead>
            <tbody>
              ${d.bom.map((b,i) => `<tr>
                <td>${i+1}</td>
                <td>${b.name||'—'}</td>
                <td class="qty">${b.qty??'—'}</td>
                <td>${b.unit||'—'}</td>
              </tr>`).join('')}
            </tbody>
          </table>
        </div>
      </div>` :
      `<div class="bom-section"><div class="section-label">Məhsullar / Xidmətlər</div><div class="no-bom">// Spesifikasiya məlumatı yoxdur</div></div>`;

    document.getElementById('result').innerHTML = `
      <div class="result-header">
        <div class="tender-title">${d.title}</div>
        <div class="status-badge ${d.expired ? 'status-expired':'status-active'}">
          ${d.expired ? '✕ Bitmişdir':'● Aktiv'}
        </div>
      </div>
      <div class="info-grid">
        <div class="info-cell full"><div class="cell-label">Təşkilat</div><div class="cell-value">${d.org}</div></div>
        <div class="info-cell"><div class="cell-label">VÖEN</div><div class="cell-value mono">${d.voen}</div></div>
        <div class="info-cell"><div class="cell-label">Təxmini məbləğ</div><div class="cell-value amount">${amt}</div></div>
        <div class="info-cell"><div class="cell-label">Başlama tarixi</div><div class="cell-value">${d.start||'—'}</div></div>
        <div class="info-cell"><div class="cell-label">Bitmə tarixi</div><div class="cell-value ${d.expired?'expired-date':''}">${d.end||'—'}</div></div>
        <div class="info-cell full"><div class="cell-label">Əlaqədar şəxs</div><div class="cell-value">${d.person}</div></div>
        <div class="info-cell"><div class="cell-label">Telefon</div><div class="cell-value">
          ${d.phone !== '—' ? `<a href="tel:${d.phone}">${d.phone}</a>` : '—'}
        </div></div>
        <div class="info-cell"><div class="cell-label">Email</div><div class="cell-value">
          ${d.email !== '—' ? `<a href="mailto:${d.email}">${d.email}</a>` : '—'}
        </div></div>
        <div class="info-cell full"><div class="cell-label">Link</div><div class="cell-value">
          <a href="${d.link}" target="_blank">${d.link} ↗</a>
        </div></div>
      </div>
      ${bomHTML}`;

    document.getElementById('result').style.display = 'block';
    document.getElementById('excelBtn').style.display = 'flex';
  }

  function setLoading(v) {
    document.getElementById('loading').style.display = v ? 'flex' : 'none';
    document.getElementById('searchBtn').disabled = v;
  }
  function showError(msg) {
    const el = document.getElementById('errorBox');
    el.textContent = '⚠ ' + msg;
    el.style.display = 'block';
  }
  function hideError() { document.getElementById('errorBox').style.display = 'none'; }

  function downloadExcel() {
    const link = document.getElementById('linkInput').value.trim();
    if (!link) return;
    window.location.href = '/api/export?link=' + encodeURIComponent(link);
  }
</script>
</body>
</html>"""

# ─────────────────────────────────────────
# ЗАПУСК
# ─────────────────────────────────────────

def open_browser():
    import time
    time.sleep(1)
    webbrowser.open('http://127.0.0.1:5000')

if __name__ == '__main__':
    print("🚀 Сервер запущен: http://127.0.0.1:5000", flush=True)
    threading.Thread(target=open_browser, daemon=True).start()
    app.run(host='0.0.0.0', debug=False, port=5000)
